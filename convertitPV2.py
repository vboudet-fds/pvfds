import pandas as pd
import fitz
import sys
import multiprocessing
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

def traiter_page(fichier, i):
    """Chaque processus ouvre une copie indépendante du PDF et traite une page."""
    try:
        doc = fitz.open(fichier)  # Ouvrir une copie du fichier PDF dans le processus
        page = doc[i]

        # Vérifier si la page contient du texte
        texte = page.get_text("text")
        if not texte.strip():
            print(f"⚠ Page {i+1} ignorée (pas de texte détecté)")
            return {}

        # Extraction des tables
        t = page.find_tables()
        if not t.tables:
            print(f"⚠ Page {i+1} ignorée (pas de tableau détecté)")
            return {}

        page_data = t[0].extract()
        entete = page_data[0]
        etudiants = {}

        for ligne in page_data[1:]:
            if ligne[0].startswith('note max'):
                break
            numero = ligne[0].split("\n")[0].split(":")[1]

            # Initialisation de l'étudiant si ce n'est pas déjà fait
            etudiant = etudiants.setdefault(numero, {'Nom\nPrénom': ligne[0].split("\n")[1]})

            for col in range(1, len(ligne)):
                if entete[col]:
                    if not ligne[col].startswith('Résultat'):
                        note = ligne[col].split("\n")[0]
                        if not note.startswith("AB") and note != "" and note != "NACQ" and note != "DIS":
                            note = float(note.split(" ")[-1])
                            etudiant[entete[col]]=note
                        else:
                            etudiant[entete[col]]=note
                    else:
                        note = ligne[col].split("\n")[0]
                        resultat = ligne[col].split("\n")[1]
                        if not note.startswith("AB") and note != "":
                            note = float(note.split(" ")[-1])
                        else:
                            note = 0
                        etudiant['Moyenne'] = note
                        etudiant['Résultat'] = resultat.split(" ")[0]
        return etudiants

    except Exception as e:
        print(f"⚠ Erreur sur la page {i+1} : {e}")  # Afficher l'erreur sans bloquer le programme
        return {}

def merge_etudiants(etudiants_list):
    """Fusionne les données de plusieurs pages pour chaque étudiant."""
    final_etudiants = {}

    for etudiants in etudiants_list:
        for numero, data in etudiants.items():
            if numero not in final_etudiants:
                final_etudiants[numero] = data
            else:
                # Fusionner les notes sans écraser
                for key, value in data.items():
                    if key in final_etudiants[numero]:
                        # Si la clé existe déjà, on écrase la valeur si ce n'est pas une liste
                        if isinstance(final_etudiants[numero][key], list):
                            final_etudiants[numero][key].append(value)
                        else:
                            final_etudiants[numero][key] = value
                    else:
                        final_etudiants[numero][key] = value

    # Moyenne recalculée si nécessaire
    for numero, data in final_etudiants.items():
        if 'Moyenne' in data and isinstance(data['Moyenne'], list):
            data['Moyenne'] = sum(data['Moyenne']) / len(data['Moyenne'])

    return final_etudiants

def export(fichier, df, simple=""):
    out=fichier.replace(".pdf", simple+".xlsx")
    df.to_excel(out,  engine="openpyxl")
    wb = load_workbook(out)
    ws = wb.active
    row_fill1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Blanc
    row_fill2 = PatternFill(start_color="FFEFD5", end_color="FFEFD5", fill_type="solid")  # Orange clair

    header_alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  

    couleurs = {
        "I": "003050",
        "C": "C20E1A",
        "E": "868686",
        "P": "612978",
        "M": "8A2F84",
        "X": "B39CC8",
        "T": "B0D2BE",
        "L": "009CDD",
        "B": "A6C236",
        "V": "CE5C37",

        "default": "999999"
    }
    polices = {
        "I": "FFFFFF",  # blanc sur fond #003050
        "C": "FFFFFF",  # blanc sur fond #C20E1A (très foncé)
        "E": "FFFFFF",  # blanc sur fond #868686 (gris foncé)
        "P": "FFFFFF",  # blanc sur fond #612978
        "M": "FFFFFF",  # blanc sur fond #8A2F84
        "X": "000000",  # noir sur fond #B39CC8 (clair)
        "T": "000000",  # noir sur fond #B0D2BE (clair)
        "L": "FFFFFF",  # blanc sur fond #009CDD
        "V": "000000",  # noir sur fond #A6C236 (clair)
        "B": "000000",  # noir sur fond #A6C236 (clair)
        "default": "FFFFFF"  # noir sur fond #DC972A (clair)
    }

    for cell in ws[1]:  
        if isinstance(cell.value, str) and len(cell.value)>6 and cell.value[6] in "ICEPMXTLVB":
            cell.fill = PatternFill(start_color=couleurs[cell.value[6]], end_color=couleurs[cell.value[6]], fill_type="solid")
            cell.font = Font(color=polices[cell.value[6]], bold=True)
        else:
            cell.fill = PatternFill(start_color=couleurs["default"], end_color=couleurs["default"], fill_type="solid")
            cell.font = Font(color=polices["default"], bold=True)
        cell.alignment = header_alignment

    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = row_fill1 if row_index % 2 == 0 else row_fill2
        for cell in row:
            cell.fill = fill
    ws.row_dimensions[1].height = 60
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):  # Colonne des moyennes
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value < 10:
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Rouge
                    cell.font = Font(bold=True)
                else:
                    cell.fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")  # Vert
                    cell.font = Font(bold=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=ws.max_column):  # Colonne des moyennes
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value < 10:
                    cell.font = Font(color="B22222")  # Rouge
                else:
                    cell.font = Font(color="228B22")   # Vert

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Obtenir la lettre de la colonne
        for cell in col[2:]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4  # Ajustement dynamique

    wb.save(out)

def convertit(fichier, progress_queue=None):
    doc = fitz.open(fichier)
    
    with multiprocessing.Pool(processes=multiprocessing.cpu_count()) as pool:
        results = pool.starmap(traiter_page, [(fichier, i) for i in range(1, len(doc) - 1)])

    etudiants = merge_etudiants(results)

    df = pd.DataFrame.from_dict(data=etudiants, orient='index')
    df2 = pd.DataFrame.from_dict(data=etudiants, orient='index')
    df2 = df2.rename(columns={col: col.split()[0] for col in df.columns})
    colonnes_a_garder = list(df2.columns[:3]) + [col for col in df2.columns[3:] 
                                         if len(str(col)) == 7 or 
                                            str(col)[-1].isdigit() or (len(str(col))>6) and col[6]=='L']

    # Filtrer le DataFrame
    df_simple = df2[colonnes_a_garder]
    export(fichier, df)
    export(fichier, df_simple,"-simple")
    return df

if __name__ == "__main__":
    if len(sys.argv) > 1:
        convertit(sys.argv[1])
    else:
        print("Usage: convertitPV fichier.pdf")
        print("Attention: l'information des UEs acquises antérieurement disparait.")


